"""
Created on Aug 11, 2021

@author: Wyatt Meehan
"""
import datetime
import re
import sys

import openpyxl


class Summary:
    def __init__(self, date, calls, aa_thirty, fcr, dsat, csat):
        self.date = date
        self.calls = calls
        self.aa_thirty = aa_thirty
        self.fcr = fcr
        self.dsat = dsat
        self.csat = csat


def get_month_str(mon):
    if mon == 1:
        return 'january'
    elif mon == 2:
        return 'february'
    elif mon == 3:
        return 'march'
    elif mon == 4:
        return 'april'
    elif mon == 5:
        return 'may'
    elif mon == 6:
        return 'june'
    elif mon == 7:
        return 'july'
    elif mon == 8:
        return 'august'
    elif mon == 9:
        return 'september'
    elif mon == 10:
        return 'october'
    elif mon == 11:
        return 'november'
    elif mon == 12:
        return 'december'
    else:
        pass


def read_workbook(path):
    try:
        workbook = openpyxl.load_workbook(path, data_only=True)
        return workbook
    except FileNotFoundError:
        print("Unable to load WorkBook, Please ensure path in main method is correct!")
        sys.exit("Program ending")
    except openpyxl.utils.exceptions.InvalidFileException:
        print("Unable to load workbook, Please ensure path in main method is correct!")
        sys.exit("Program ending")


def read_worksheet(w_b, path):
    try:
        worksheet1 = w_b[path]
        return worksheet1
    except FileNotFoundError:
        print("Unable to load WorkSheet, Please ensure path in main method is correct!")
        sys.exit("Program ending")
    except KeyError:
        print("Unable to load WorkSheet, Please ensure path in main method is correct!")
        sys.exit("Program ending")


def read_summary(path, worksheet):
    workbook = read_workbook(path)
    worksheet1 = read_worksheet(workbook, worksheet)
    pattern = "expedia_report_monthly_(.*?).xlsx"
    month_year = re.search(pattern, path).group(1)
    split_str = month_year.partition('_')
    month_str = split_str[0]
    year_str = split_str[2]
    month_list = []
    for row in worksheet1.iter_rows(worksheet1.min_row, worksheet1.max_row):
        for cell in row:
            if cell.value is not None and cell.value and isinstance(cell.value, datetime.datetime):
                month = Summary(cell.value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value)
                month_list.append(month)
    for month_x in month_list:
        if month_str == get_month_str(month_x.date.month) and str(year_str).upper() == str(month_x.date.year).upper():
            print("\nEg for " + month_str.capitalize() + ":\n")
            print("Calls Offered: {} \nAbandon after 30s:"
                  " {:.2f}% \nFCR: {:.2f}% \nDSAT: {:.2f}% \nCSAT: {:.2f}%".format(month_x.calls,
                                                                                   month_x.aa_thirty * 100,
                                                                                   month_x.fcr * 100,
                                                                                   month_x.dsat * 100,
                                                                                   month_x.csat * 100))


def read_voc(path, worksheet):
    workbook = read_workbook(path)
    worksheet2 = read_worksheet(workbook, worksheet)
    pattern = "expedia_report_monthly_(.*?).xlsx"
    month_year = re.search(pattern, path).group(1)
    split_str = month_year.partition('_')
    month_str = split_str[0][:3].capitalize()
    year_str = split_str[2]

    for column in worksheet2.iter_cols(worksheet2.min_column, worksheet2.max_column):
        for cell in column:
            if cell.value is not None and cell.value and isinstance(cell.value, datetime.datetime):
                if get_month_str(cell.value.month)[:3].capitalize() == month_str and str(cell.value.year) == year_str:
                    print(
                        "\nBase Size: {}\nPromoters (Recommend Score 9 to 10):  {}\nPassives (Recommend Score 7 to "
                        "8): {}\nDetractors (recommend Score 0 to 6): {}\nAARP Total: {:.2f}\nSat with Agent %: {:.2f}"
                        "\nDSat "
                        "with Agent %: {:.2f} ".format(column[2].value,
                                                       column[3].value,
                                                       column[5].value,
                                                       column[7].value,
                                                       column[12].value * 100,
                                                       column[15].value * 100,
                                                       column[18].value * 100))
                    print("\nIn net Promoter Score: ")
                    if column[3].value >= 200:
                        print("Promoters:  Good!  ", column[3].value)
                    else:
                        print("Promoters:  Bad!  ", column[3].value)
                    if column[5].value >= 100:
                        print("Passives:   Good!  ", column[5].value)
                    else:
                        print("Passives:   Bad!  ", column[5].value)
                    if column[7].value >= 100:
                        print("Detractors: Good!  ", column[7].value)
                    else:
                        print("Detractors: Bad!  ", column[7].value)


if __name__ == '__main__':
    excel_path = ("C:\\Users\\meeha\\git\\PythonProjects\\PythonProjects\\.pydevproject\\PythonProjects\\ss\\python"
                  "\\excel\\expedia_report_monthly_january_2018.xlsx")
    read_summary(excel_path, "Summary Rolling MoM")
    read_voc(excel_path, "VOC Rolling MoM")
